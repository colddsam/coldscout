"""
Client proposal spreadsheet generator.
Creates a visually formatted Excel workbook for the client that contains:
  Sheet 1 — ROI Projection  (editable assumptions, auto-calculated returns)
  Sheet 2 — Competitor Snapshot  (their gap vs competitors)
  Sheet 3 — Project Roadmap  (timeline with deliverables)

Designed to be impressive, interactive, and personalised per lead.
Brand identity: black & white with shades of gray.
"""
import os
from datetime import date, timedelta
from typing import List, Optional, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from loguru import logger
from app.config import get_settings

settings = get_settings()

# ── Brand colours (openpyxl uses ARGB hex — Black & White theme) ──────────────
BLACK      = "000000"
DARK_GRAY  = "333333"
MID_GRAY   = "666666"
GRAY_500   = "999999"
GRAY_300   = "CCCCCC"
BORDER_CLR = "EAEAEA"
WHITE      = "FFFFFF"
LIGHT_BG   = "F5F5F5"
LIGHTEST   = "FAFAFA"

# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=BLACK, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_bottom(color=GRAY_300) -> Border:
    side = Side(style="thin", color=color)
    return Border(bottom=side)

def _thick_bottom(color=BLACK) -> Border:
    side = Side(style="medium", color=color)
    return Border(bottom=side)

def _cell(ws, row, col, value, bold=False, size=11, color=BLACK,
          fill=None, align="left", valign="center", wrap=False, italic=False,
          number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold, size, color, italic)
    c.alignment = _align(align, valign, wrap)
    if fill:
        c.fill = _fill(fill)
    if number_format:
        c.number_format = number_format
    return c

def _merge_title(ws, row, start_col, end_col, value, bg=BLACK, fg=WHITE,
                 size=14, bold=True, height=30):
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row,   end_column=end_col
    )
    c = ws.cell(row=row, column=start_col, value=value)
    c.font      = _font(bold, size, fg)
    c.fill      = _fill(bg)
    c.alignment = _align("center", "center")
    ws.row_dimensions[row].height = height
    return c

def _section_header(ws, row, start_col, end_col, value, bg=BLACK, fg=WHITE,
                    size=10):
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row,   end_column=end_col
    )
    c = ws.cell(row=row, column=start_col, value=value)
    c.font      = _font(True, size, fg)
    c.fill      = _fill(bg)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = 22
    return c

def _col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ── Sheet 1: ROI Projection ───────────────────────────────────────────────────

def _build_roi_sheet(ws, business_name: str, category: str, city: str,
                     rating: float, review_count: int):
    ws.sheet_view.showGridLines = False

    # Title block
    _merge_title(ws, 1, 1, 8,
                 f"Digital ROI Projection  —  {business_name}",
                 BLACK, WHITE, 16, True, 36)

    ws.merge_cells("A2:H2")
    sub = ws.cell(row=2, column=1,
                  value=f"{category}  |  {city}  |  {rating}★  ({review_count} reviews)  |  Prepared: {date.today().strftime('%d %B %Y')}")
    sub.font      = _font(False, 9, GRAY_300)
    sub.fill      = _fill(BLACK)
    sub.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 18

    # ── Assumptions block ─────────────────────────────────────────────────────
    _section_header(ws, 4, 1, 4, "  📊  KEY ASSUMPTIONS  (gray = change these)", DARK_GRAY, WHITE, 9)

    assumptions = [
        ("Current monthly website visitors",   "B5",  0,      "#,##0"),
        ("Average customer value (₹)",         "B6",  3500,   "₹#,##0"),
        ("Current monthly enquiries (calls / emails)", "B7", 8, "#,##0"),
        ("Expected traffic uplift after launch (%)",   "B8", 180, "0%"),
        ("Enquiry conversion rate (%)",         "B9",  12,     "0%"),
        ("Project investment (₹)",             "B10", 35000,  "₹#,##0"),
    ]
    labels_col = ["Assumption", "", "", ""]
    for i, (label, _, default, fmt) in enumerate(assumptions):
        row = 5 + i
        _cell(ws, row, 1, label, size=9, color=MID_GRAY)
        c = ws.cell(row=row, column=2, value=default)
        c.font         = _font(True, 10, BLACK)
        c.alignment    = _align("center", "center")
        c.number_format = fmt
        c.fill         = _fill(LIGHT_BG)
        ws.row_dimensions[row].height = 18

    # ── Projected results ─────────────────────────────────────────────────────
    _section_header(ws, 12, 1, 4, "  📈  PROJECTED RESULTS  (auto-calculated)", DARK_GRAY, WHITE, 9)

    results = [
        ("Projected monthly visitors",   "=B5*(1+B8)",       "#,##0"),
        ("Additional visitors / month",  "=B13-B5",          "#,##0"),
        ("New monthly enquiries",        "=ROUND(B13*B9,0)", "#,##0"),
        ("Additional revenue / month",   "=B15*B6",          "₹#,##0"),
        ("Annual revenue uplift",        "=B16*12",          "₹#,##0"),
        ("Payback period (months)",      "=IFERROR(ROUND(B10/B16,1),\"N/A\")", "0.0"),
        ("12-month ROI",                 "=IFERROR(ROUND((B17-B10)/B10*100,0)&\"%\",\"N/A\")", "@"),
    ]
    for i, (label, formula, fmt) in enumerate(results):
        row = 13 + i
        _cell(ws, row, 1, label, size=9, color=MID_GRAY)
        c = ws.cell(row=row, column=2, value=formula)
        c.font          = _font(True, 10, BLACK)
        c.alignment     = _align("center", "center")
        c.number_format = fmt
        c.fill          = _fill(LIGHTEST)
        ws.row_dimensions[row].height = 20

    # Thick border under payback
    for col in range(1, 5):
        ws.cell(row=19, column=col).border = _thick_bottom()

    # ── Monthly revenue build chart ───────────────────────────────────────────
    # Build a small data table for the chart (months 1–12)
    ws.cell(row=22, column=1, value="Month").font = _font(True, 9, MID_GRAY)
    ws.cell(row=22, column=2, value="Cumulative Revenue Uplift (₹)").font = _font(True, 9, MID_GRAY)
    ws.cell(row=22, column=3, value="Investment").font = _font(True, 9, MID_GRAY)

    for m in range(1, 13):
        ws.cell(row=22 + m, column=1, value=f"Month {m}")
        ws.cell(row=22 + m, column=2, value=f"=B16*{m}")
        ws.cell(row=22 + m, column=2).number_format = "₹#,##0"
        ws.cell(row=22 + m, column=3, value="=B10")
        ws.cell(row=22 + m, column=3).number_format = "₹#,##0"

    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Cumulative Revenue vs Investment Over 12 Months"
    chart.y_axis.title = "₹ Value"
    chart.x_axis.title = "Month"
    chart.style  = 10
    chart.width  = 16
    chart.height = 10
    chart.grouping = "clustered"

    data   = Reference(ws, min_col=2, max_col=3, min_row=22, max_row=34)
    cats   = Reference(ws, min_col=1, min_row=23, max_row=34)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = BLACK
    chart.series[1].graphicalProperties.solidFill = GRAY_300
    ws.add_chart(chart, "E4")

    _col_widths(ws, {"A": 36, "B": 20, "C": 22, "D": 4,
                     "E": 14, "F": 14, "G": 14, "H": 14})


# ── Sheet 2: Competitor Snapshot ──────────────────────────────────────────────

def _build_competitor_sheet(ws, business_name: str, category: str):
    ws.sheet_view.showGridLines = False

    _merge_title(ws, 1, 1, 7, f"Competitor Digital Snapshot  —  {category}", BLACK, WHITE, 14)
    ws.merge_cells("A2:G2")
    sub = ws.cell(row=2, column=1,
                  value="How does your digital presence compare to top competitors in your category?")
    sub.font      = _font(False, 9, WHITE)
    sub.fill      = _fill(BLACK)
    sub.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 18

    # Headers — do NOT merge this row; write each cell individually
    headers = ["Business", "Website", "SSL (HTTPS)", "Mobile Friendly",
               "Social Media", "Online Booking", "Digital Score"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font      = _font(True, 9, WHITE)
        c.fill      = _fill(DARK_GRAY)
        c.alignment = _align("center", "center")
    ws.row_dimensions[4].height = 22

    # Data rows
    rows = [
        [f"⭐ {business_name} (You)", "❌ None", "—", "—", "—", "—", "Low"],
        ["Competitor A", "✅ Yes", "✅ Yes", "✅ Yes", "✅ Yes", "✅ Yes", "High"],
        ["Competitor B", "✅ Yes", "✅ Yes", "❌ No",  "✅ Yes", "❌ No",  "Medium"],
        ["Competitor C", "✅ Yes", "❌ No",  "✅ Yes", "❌ No",  "❌ No",  "Medium"],
        ["Category Average", "83%", "75%", "80%", "70%", "45%", "Medium"],
    ]
    score_fills = {"High": DARK_GRAY, "Medium": MID_GRAY, "Low": BLACK}
    for ri, row_data in enumerate(rows):
        row_num   = 5 + ri
        row_fill  = LIGHT_BG if ri == 0 else (LIGHTEST if ri % 2 == 0 else WHITE)
        bold_row  = ri == 0
        for ci, val in enumerate(row_data):
            c = ws.cell(row=row_num, column=ci + 1, value=val)
            c.font      = _font(bold_row, 9, BLACK)
            c.fill      = _fill(row_fill)
            c.alignment = _align("center" if ci > 0 else "left", "center")
            if ci == 6:  # Score column
                score_color = score_fills.get(val, GRAY_300)
                c.fill = _fill(score_color)
                c.font = _font(True, 9, WHITE)
        ws.row_dimensions[row_num].height = 20

    # Note
    ws.merge_cells("A11:G11")
    note = ws.cell(row=11, column=1,
                   value="💡  Note: Competitor data is illustrative. After your digital upgrade, your score moves from Low → High.")
    note.font      = _font(False, 8, MID_GRAY, italic=True)
    note.alignment = _align("left", "center")
    note.fill      = _fill(LIGHT_BG)
    ws.row_dimensions[11].height = 20

    _col_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 18,
                     "E": 16, "F": 18, "G": 16})


# ── Sheet 3: Project Roadmap ──────────────────────────────────────────────────

def _build_roadmap_sheet(ws, benefits: List[str]):
    ws.sheet_view.showGridLines = False

    _merge_title(ws, 1, 1, 6, "Project Roadmap & Deliverables", BLACK, WHITE, 14)
    ws.row_dimensions[1].height = 36

    start   = date.today()
    phases  = [
        {
            "phase": "Phase 1 — Discovery & Strategy",
            "color": BLACK,
            "duration_days": 14,
            "deliverables": [
                "Brand & competitor audit report",
                "Keyword research (50+ local search terms)",
                "Site architecture & content plan",
                "Design mood board approval",
            ],
        },
        {
            "phase": "Phase 2 — Design & Development",
            "color": DARK_GRAY,
            "duration_days": 28,
            "deliverables": [
                "Mobile-first responsive website (5–8 pages)",
                "On-page SEO for all pages",
                "Contact form & lead capture setup",
                "Speed optimisation (target < 2 s load time)",
            ],
        },
        {
            "phase": "Phase 3 — Launch & Visibility",
            "color": MID_GRAY,
            "duration_days": 14,
            "deliverables": [
                "Live website deployment",
                "Google Business Profile setup / optimisation",
                "Google Analytics 4 + Search Console",
                "30-day post-launch performance check",
            ],
        },
    ]

    headers = ["Phase", "Start Date", "End Date", "Duration", "Deliverable", "Status"]
    row = 3
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font      = _font(True, 9, WHITE)
        c.fill      = _fill(DARK_GRAY)
        c.alignment = _align("center", "center")
    ws.row_dimensions[row].height = 22
    row += 1

    current_start = start
    for phase in phases:
        phase_end = current_start + timedelta(days=phase["duration_days"])
        color     = phase["color"]

        for di, deliverable in enumerate(phase["deliverables"]):
            show_phase = di == 0
            r_fill     = LIGHTEST if di % 2 == 0 else WHITE

            # Phase name cell (only first row of each phase)
            c = ws.cell(row=row, column=1,
                        value=phase["phase"] if show_phase else "")
            c.font      = _font(True, 9, WHITE if show_phase else BLACK)
            c.fill      = _fill(color if show_phase else r_fill)
            c.alignment = _align("left", "center", wrap=True)

            # Start date
            c2 = ws.cell(row=row, column=2,
                         value=current_start.strftime("%d %b %Y") if show_phase else "")
            c2.font      = _font(False, 9, MID_GRAY)
            c2.fill      = _fill(r_fill)
            c2.alignment = _align("center", "center")

            # End date
            c3 = ws.cell(row=row, column=3,
                         value=phase_end.strftime("%d %b %Y") if show_phase else "")
            c3.font      = _font(False, 9, MID_GRAY)
            c3.fill      = _fill(r_fill)
            c3.alignment = _align("center", "center")

            # Duration
            c4 = ws.cell(row=row, column=4,
                         value=f"{phase['duration_days']} days" if show_phase else "")
            c4.font      = _font(False, 9, MID_GRAY)
            c4.fill      = _fill(r_fill)
            c4.alignment = _align("center", "center")

            # Deliverable
            c5 = ws.cell(row=row, column=5, value=f"  ▸  {deliverable}")
            c5.font      = _font(False, 9, BLACK)
            c5.fill      = _fill(r_fill)
            c5.alignment = _align("left", "center")

            # Status
            c6 = ws.cell(row=row, column=6, value="Planned")
            c6.font      = _font(False, 8, MID_GRAY)
            c6.fill      = _fill(r_fill)
            c6.alignment = _align("center", "center")

            ws.row_dimensions[row].height = 20
            row += 1

        current_start = phase_end + timedelta(days=1)
        row += 1  # gap between phases

    # Benefits delivered summary
    row += 1
    _section_header(ws, row, 1, 6, "  ✅  KEY OUTCOMES THIS PROJECT DELIVERS", DARK_GRAY, WHITE, 9)
    row += 1
    for benefit in (benefits or [])[:5]:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=6
        )
        c = ws.cell(row=row, column=1, value=f"  ✓  {benefit}")
        c.font      = _font(False, 9, MID_GRAY)
        c.fill      = _fill(LIGHT_BG)
        c.alignment = _align("left", "center")
        ws.row_dimensions[row].height = 20
        row += 1

    _col_widths(ws, {"A": 32, "B": 14, "C": 14, "D": 12, "E": 44, "F": 14})


# ── Public interface ──────────────────────────────────────────────────────────

def generate_proposal_xlsx(
    business_name: str,
    category: str,
    benefits: List[str],
    output_filename: str = "proposal.xlsx",
    rating: Optional[float] = None,
    review_count: Optional[int] = None,
    city: Optional[str] = None,
) -> Optional[str]:
    """
    Generates a three-sheet Excel proposal workbook for the client.

    Sheets:
        1. ROI Projection   — editable assumptions, auto-calculated returns, chart
        2. Competitor Gap   — illustrative digital benchmarking table
        3. Project Roadmap  — phased timeline with deliverables

    Args:
        business_name:   Target business display name.
        category:        Business category (e.g. "Dentists").
        benefits:        AI-generated value propositions (3–5 items).
        output_filename: Output filename inside tmp/.
        rating:          Google rating (optional).
        review_count:    Number of Google reviews (optional).
        city:            Business city (optional).

    Returns:
        str: Path to the generated file, or None on failure.
    """
    try:
        os.makedirs(settings.ATTACHMENT_DIR, exist_ok=True)
        filepath = os.path.join(settings.ATTACHMENT_DIR, output_filename)

        wb = Workbook()

        # Sheet 1 — ROI Projection
        ws1 = wb.active
        ws1.title = "📊 ROI Projection"
        _build_roi_sheet(
            ws1,
            business_name,
            category,
            city or "Your City",
            rating or 4.2,
            review_count or 50,
        )

        # Sheet 2 — Competitor Snapshot
        ws2 = wb.create_sheet("🔍 Competitor Snapshot")
        _build_competitor_sheet(ws2, business_name, category)

        # Sheet 3 — Roadmap
        ws3 = wb.create_sheet("🗓️ Project Roadmap")
        _build_roadmap_sheet(ws3, benefits or [])

        # Tab colours (grayscale)
        ws1.sheet_properties.tabColor = BLACK
        ws2.sheet_properties.tabColor = DARK_GRAY
        ws3.sheet_properties.tabColor = MID_GRAY

        wb.save(filepath)
        logger.info(f"Proposal xlsx generated: {filepath}")
        return filepath

    except Exception as e:
        logger.exception(f"Failed to generate xlsx proposal for {business_name}: {e}")
        return None