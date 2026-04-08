"""
Excel report generation module.
Utilizes the openpyxl library to systematically assemble quantitative metrics
and lead status details into a formatted administrative spreadsheet artifact.
Brand identity: black & white with shades of gray.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
import os
from typing import List, Dict, Any
from app.config import get_settings

settings = get_settings()

def generate_daily_report_excel(report_data: Dict[str, Any], leads: List[Dict[str, Any]], output_date: date) -> str:
    """
    Compiles daily performance metrics and individual lead properties into 
    a multi-sheet formatted Excel workbook.
    
    Args:
        report_data (Dict[str, Any]): Dictionary containing foundational daily metrics.
        leads (List[Dict[str, Any]]): Structurally formatted collection of recent lead data.
        output_date (date): The contextual date associated with the report.
        
    Returns:
        str: Expected file path to the synchronously generated Excel document.
    """
    wb = openpyxl.Workbook()
    
    # ── Brand styles ──────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")
    
    data_font = Font(name="Calibri", size=11, color="000000")
    data_font_muted = Font(name="Calibri", size=11, color="666666")
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    border_bottom = Border(bottom=Side(style="thin", color="EAEAEA"))
    
    # ── Sheet 1: Daily Summary ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1.sheet_view.showGridLines = False
    
    # Title row
    ws1.merge_cells("A1:B1")
    title_cell = ws1.cell(row=1, column=1, value=f"Cold Scout  ·  Daily Report  ·  {output_date}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32
    
    # Column headers
    ws1.cell(row=3, column=1, value="Metric").font = header_font
    ws1.cell(row=3, column=1).fill = header_fill
    ws1.cell(row=3, column=1).alignment = header_align
    ws1.cell(row=3, column=2, value="Value").font = header_font
    ws1.cell(row=3, column=2).fill = header_fill
    ws1.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 24
    
    metrics = [
        ("Total leads discovered", report_data.get('leads_discovered', 0)),
        ("Qualified leads", report_data.get('leads_qualified', 0)),
        ("Emails sent", report_data.get('emails_sent', 0)),
        ("Emails opened", report_data.get('emails_opened', 0)),
        ("Links clicked", report_data.get('links_clicked', 0)),
        ("Replies received", report_data.get('replies_received', 0)),
    ]
    
    for i, (label, value) in enumerate(metrics):
        row = 4 + i
        c1 = ws1.cell(row=row, column=1, value=label)
        c1.font = data_font
        c1.border = border_bottom
        c2 = ws1.cell(row=row, column=2, value=value)
        c2.font = Font(name="Calibri", bold=True, size=12, color="000000")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = border_bottom
        if i % 2 == 0:
            c1.fill = alt_fill
            c2.fill = alt_fill
        ws1.row_dimensions[row].height = 22
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 16
    
    # ── Sheet 2: Lead Details ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Lead Details")
    ws2.sheet_view.showGridLines = False
    
    headers = ["Business", "Category", "Country", "Region", "Location", "Email Sent", "Opened", "Clicked", "Replied", "Status", "Phone", "Google Maps"]
    for i, h in enumerate(headers):
        c = ws2.cell(row=1, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for ri, lead in enumerate(leads):
        row = 2 + ri
        values = [
            lead.get("business_name"),
            lead.get("category"),
            lead.get("country", ""),
            lead.get("region", ""),
            lead.get("city"),
            "Yes" if lead.get("email_sent_at") else "No",
            "Yes" if lead.get("first_opened_at") else "No",
            "Yes" if lead.get("first_clicked_at") else "No",
            "Yes" if lead.get("first_replied_at") else "No",
            lead.get("status"),
            lead.get("phone"),
            lead.get("google_maps_url")
        ]
        for ci, val in enumerate(values):
            c = ws2.cell(row=row, column=ci + 1, value=val)
            c.font = data_font if ci == 0 else data_font_muted
            c.alignment = Alignment(horizontal="center" if ci > 0 else "left", vertical="center")
            c.border = border_bottom
            if ri % 2 == 0:
                c.fill = alt_fill
        ws2.row_dimensions[row].height = 20

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['L'].width = 36
    
    # Tab colors (grayscale)
    ws1.sheet_properties.tabColor = "000000"
    ws2.sheet_properties.tabColor = "333333"
        
    os.makedirs(settings.ATTACHMENT_DIR, exist_ok=True)
    filename = f"ColdScout_Report_{output_date.strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(settings.ATTACHMENT_DIR, filename)
    wb.save(filepath)
    
    return filepath
